from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Grades.xlsx')
ws = wb.active
# change the value
# ws['A2'].value = 'Sam'


# Creating a new sheet
# wb.create_sheet('new sheet')

# print(wb.sheetnames)

# Merging cells
# ws.merge_cells('A1:D1')

# Unmerging cells
# ws.unmerge_cells('A1:D1')

# Insert an empty rows (after!)
# ws.insert_rows(7)


# Delete (after!)
# ws.delete_rows(7)

# Copying and moving cells
# ws.move_range('C1:D11', rows=2, cols=2)


for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        # print(ws[char + str(row)].value)
        ws[char + str(row)] = char + str(row)


# Saving
wb.save('Grades.xlsx')
